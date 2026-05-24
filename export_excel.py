import pandas as pd
import os
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

DB_URL = "postgresql://postgres:root@localhost:5432/Stock_Index"


def fetch_data(engine):
    monthly_summary = pd.read_sql("""
        SELECT
            i.index_name,
            DATE_TRUNC('month', ti.date)::DATE            AS month,
            ROUND(AVG(ti.volatility_30d)::NUMERIC, 5)     AS avg_volatility,
            ROUND(AVG(ti.daily_return)::NUMERIC, 5)        AS avg_daily_return,
            ROUND(SUM(ti.daily_return)::NUMERIC, 4)        AS total_monthly_return
        FROM technical_indicators ti
        JOIN indexes i ON i.index_id = ti.index_id
        GROUP BY i.index_name, month
        ORDER BY month, i.index_name
    """, engine)

    rsi_signals = pd.read_sql("""
        SELECT
            i.index_name,
            ti.date,
            ROUND(ti.rsi::NUMERIC, 2) AS rsi,
            ti.rsi_signal
        FROM technical_indicators ti
        JOIN indexes i ON i.index_id = ti.index_id
        WHERE ti.rsi_signal IN ('Overbought', 'Oversold')
        ORDER BY ti.date DESC
        LIMIT 100
    """, engine)

    correlation = pd.read_sql("""
        SELECT
            ia.index_name                                            AS index_a,
            ib.index_name                                            AS index_b,
            ROUND(CORR(a.daily_return, b.daily_return)::NUMERIC, 4) AS correlation
        FROM technical_indicators a
        JOIN technical_indicators b ON a.date = b.date
        JOIN indexes ia              ON ia.index_id = a.index_id
        JOIN indexes ib              ON ib.index_id = b.index_id
        WHERE ia.index_name < ib.index_name
        GROUP BY ia.index_name, ib.index_name
        ORDER BY correlation DESC
    """, engine)

    drawdowns = pd.read_sql("""
        WITH running_max AS (
            SELECT
                i.index_name,
                o.date,
                o.close,
                MAX(o.close) OVER (
                    PARTITION BY o.index_id
                    ORDER BY o.date
                    ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ) AS peak_close
            FROM daily_ohlcv o
            JOIN indexes i ON i.index_id = o.index_id
        )
        SELECT
            index_name,
            date,
            ROUND(close::NUMERIC, 2)      AS close,
            ROUND(peak_close::NUMERIC, 2) AS peak_close,
            ROUND(((close - peak_close) / peak_close * 100)::NUMERIC, 2) AS drawdown_pct
        FROM running_max
        ORDER BY drawdown_pct ASC
        LIMIT 50
    """, engine)

    return monthly_summary, rsi_signals, correlation, drawdowns


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(ws, min_width=12, max_width=28):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=min_width)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_color_scale(ws, data_range):
    rule = ColorScaleRule(
        start_type="min", start_color="DDEEFF",
        end_type="max",   end_color="08306B"
    )
    ws.conditional_formatting.add(data_range, rule)


def build_excel():
    engine = create_engine(DB_URL)
    monthly_summary, rsi_signals, correlation, drawdowns = fetch_data(engine)

    # Pivot for volatility heatmap
    heatmap = monthly_summary.pivot(
        index="index_name", columns="month", values="avg_volatility"
    )
    heatmap.columns = [str(c)[:7] for c in heatmap.columns]

    output_path = "data/index_market_report.xlsx"
    os.makedirs("data", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        monthly_summary.to_excel(writer, sheet_name="Monthly Summary",    index=False)
        heatmap.to_excel(        writer, sheet_name="Volatility Heatmap")
        rsi_signals.to_excel(    writer, sheet_name="RSI Signals",        index=False)
        correlation.to_excel(    writer, sheet_name="Pair Correlation",   index=False)
        drawdowns.to_excel(      writer, sheet_name="Drawdown Analysis",  index=False)

    # Post-process with openpyxl
    wb = load_workbook(output_path)

    # Monthly Summary
    ws = wb["Monthly Summary"]
    style_header(ws, 1, 5)
    autofit_columns(ws)

    # Volatility Heatmap — color scale on data cells
    ws2 = wb["Volatility Heatmap"]
    style_header(ws2, 1, ws2.max_column)
    if ws2.max_row > 1 and ws2.max_column > 1:
        data_range = f"B2:{get_column_letter(ws2.max_column)}{ws2.max_row}"
        apply_color_scale(ws2, data_range)
    autofit_columns(ws2)

    # RSI Signals — red for overbought, green for oversold
    ws3 = wb["RSI Signals"]
    style_header(ws3, 1, 4)
    red_fill   = PatternFill("solid", fgColor="FFDDC1")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        signal = row[3].value
        fill = red_fill if signal == "Overbought" else (
               green_fill if signal == "Oversold" else None)
        if fill:
            for cell in row:
                cell.fill = fill
    autofit_columns(ws3)

    # Correlation
    ws4 = wb["Pair Correlation"]
    style_header(ws4, 1, 3)
    autofit_columns(ws4)

    # Drawdown Analysis
    ws5 = wb["Drawdown Analysis"]
    style_header(ws5, 1, 5)
    orange_fill = PatternFill("solid", fgColor="FFE0CC")
    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row):
        dd = row[4].value
        if dd is not None and float(dd) < -10:
            for cell in row:
                cell.fill = orange_fill
    autofit_columns(ws5)

    wb.save(output_path)
    print(f"Excel report saved → {output_path}")


if __name__ == "__main__":
    build_excel()
